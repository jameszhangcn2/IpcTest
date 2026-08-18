# conftest.py
import pytest
from pathlib import Path
from utils.camera_recorder import CameraRecorder
from utils.camera_picture import CameraPicture
from utils.canoe_api import CanoeApi
from utils.serial_kl15 import SerialKL15

import pytest
import pythoncom
import win32com.client
import time
import datetime
import cv2
import os
import subprocess
import numpy as np
from skimage.metrics import structural_similarity
from pathlib import Path
from typing import Optional
import threading
import sys
from datetime import datetime

from openpyxl import Workbook, load_workbook

BASE_DIR = Path(__file__).resolve().parent
VIDEO_DIR = BASE_DIR / "testlogs"
PICTURE_DIR = BASE_DIR / "testlogs"
CAMERA_INDEX_PICTURE = 0   # 第0号摄像头，多摄像头可改成1,2
CAMERA_INDEX_VIDEO = 2
REC_FPS = 12
KL15COM_PORT="COM6"
# ----------------------配置修改成你自己的路径----------------------
#CANOE_CFG = str(BASE_DIR / "CANoe" / "test.cfg")# CANoe工程cfg绝对路径
CANOE_CFG = str(BASE_DIR / "CANoe" / "332Replay.cfg")# CANoe工程cfg绝对路径
#TEMPLATE_IMG = str(BASE_DIR / "CANoe" / "demo.jpg") # HMI参考模板图
# ----------------------------------------------------------------

# 1. 注册自定义命令行参数 --log-root
def pytest_addoption(parser):
    parser.addoption(
        "--log-root",
        action="store",
        default="./testlogs",  # 默认日志根目录
        help="顶层日志根目录路径，例：--log-root=C:/AutoTest/logs"
    )

# 2. session级全局fixture：拿到根目录
@pytest.fixture(scope="session")
def log_root(pytestconfig):
    root_path = Path(pytestconfig.getoption("--log-root")).resolve()
    root_path.mkdir(parents=True, exist_ok=True)
    return root_path

# 3. session日志总目录（带时间戳，所有用例共用）
@pytest.fixture(scope="session")
def dir_session(log_root):
    session_dir = log_root
    session_dir.mkdir(parents=True, exist_ok=True)
    print(f"✅ Session全局日志目录：{session_dir}")
    return session_dir
    
@pytest.fixture(scope="session")
def cam_recorder():
    rec = CameraRecorder(save_root=VIDEO_DIR, camera_id=CAMERA_INDEX_VIDEO)
    yield rec

@pytest.fixture(scope="session")
def cam_picture():
    rec = CameraPicture(save_root=PICTURE_DIR, camera_id=CAMERA_INDEX_PICTURE)
    yield rec
    
@pytest.fixture(scope="function", autouse=True)
def auto_camera_record(cam_recorder, request, case_logger_dir, case_logger):
    case_name = request.node.name
    stop_event = threading.Event()

    # 开启摄像头录像
    cam_recorder.start_record(case_name, case_logger_dir)

    # 后台线程持续取帧
    def frame_loop():
        interval = 1.0 / cam_recorder.fps
        while not stop_event.is_set():
            cam_recorder.grab_frame()
            time.sleep(interval)

    rec_thread = threading.Thread(target=frame_loop, daemon=True)
    rec_thread.start()

    yield  # 执行测试用例

    # 用例结束，停止录制
    stop_event.set()
    rec_thread.join()
    cam_recorder.stop_record()


# =========可选：只失败保留视频，成功删除=========
@pytest.hookimpl(tryfirst=True, hookwrapper=True)
def pytest_runtest_makereport(item, call):
    outcome = yield
    rep = outcome.get_result()
    setattr(item, "rep_" + rep.when, rep)


def kill_canoe_process():
    try:
        subprocess.run(["taskkill", "/F", "/IM", "CANoe64.exe"], capture_output=True, shell=False)
        time.sleep(1.5)
    except Exception:
        pass
        
@pytest.fixture(scope="session")
def canoe_api(kl15):
    """
    session级别fixture：所有用例执行前启动CANoe，全部跑完关闭CANoe
    yield之前=前置；yield之后=后置清理，就算用例失败也执行
    """
    canoeApi = CanoeApi()
    kill_canoe_process()
    time.sleep(5)
    print("\n==== 连接CANoe COM ====")
    canoeApi.app = win32com.client.Dispatch("CANoe.Application")
    canoeApi.app.Visible = True   # True显示CANoe窗口；False后台运行
    print("\n Current COM CANOE version:", canoeApi.app.version)
 
    assert os.path.exists(CANOE_CFG)
    time.sleep(2)
    
    
    # 打开工程配置
    canoeApi.app.Open(CANOE_CFG)
    # 确认已加载
    config = canoeApi.app.Configuration
    print(f"✅ 已加载配置: {config.Name}")  # 应该显示你工程里的配置名（如 Configuration 1）
    kl15.kl15on() #keyon KL15
    time.sleep(2)
    # 显示 Configuration 视图
    #app.ActivateView(0)
    #app.ConfigurationWindow.Visible = True
        
    measurement = canoeApi.app.Measurement
 
    # 启动测量
    #if not measurement.Running:
    #    measurement.Start()
    #    time.sleep(2)
 
    yield canoeApi   # 把canoe对象传给测试用例
 
    # --------后置清理：全部用例跑完执行--------
    print("\n==== 关闭CANoe ====")
    if measurement.Running:
        measurement.Stop()
        kl15.kl15off() #keyon KL15
    canoeApi.app.Quit()
 

@pytest.fixture(scope="session")
def kl15():
    """
    session级别fixture：所有用例执行前启动CANoe，全部跑完关闭CANoe
    yield之前=前置；yield之后=后置清理，就算用例失败也执行
    """
    kl15 = SerialKL15(KL15COM_PORT)
    
    assert kl15.isKl15Open()
    time.sleep(2)
    
    yield kl15   # 把canoe对象传给测试用例
 
    # --------后置清理：全部用例跑完执行--------
    print("\n==== 关闭KL15 serial port ====")
    kl15.ser.close()
    

@pytest.fixture(scope="function")
def case_logger_dir(request, dir_session):
    # 获取当前用例完整名称，处理非法文件名字符
    case_name = request.node.nodeid.replace("/", "_").replace("\\", "_").replace(":", "_")
    log_dir = Path(dir_session)
    os.makedirs(log_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = os.path.join(log_dir, f"{case_name}_{timestamp}")

    print(f"===== this case log dir：{log_path} =====")
    yield log_path
    print(f"===== this case log end：{log_path} =====")

@pytest.fixture(scope="function")
def case_logger(case_logger_dir, request):
    log_dir = Path(case_logger_dir)
    print("case_logger 文件夹是否存在：", log_dir.exists())
    os.makedirs(log_dir, exist_ok=True)
    time.sleep(1)
    print("case_logger 文件夹是否存在：", log_dir.exists())
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = os.path.join(case_logger_dir, f"ScriptLog_{timestamp}.log")
    log_fp = open(log_path, "w", encoding="utf-8")

    # 保存原始输出流
    old_stdout = sys.stdout
    old_stderr = sys.stderr

    class DualOutput:
        def write(self, text):
            old_stdout.write(text)
            log_fp.write(text)
            log_fp.flush()
        def flush(self):
            old_stdout.flush()

    stream = DualOutput()
    sys.stdout = stream
    sys.stderr = stream

    print(f"===== 测试用例开始：{request.node.nodeid} =====")
    yield log_path

    # 用例执行完成后恢复流，关闭文件
    print(f"===== 测试用例结束：{request.node.nodeid} =====")
    sys.stdout = old_stdout
    sys.stderr = old_stderr
    log_fp.close()
    print(f"本条用例日志已保存：{log_path}")

# 全局变量：excel路径，会话启动时初始化
excel_save_path: Path = None


def init_excel_file(file_path: Path):
    """创建Excel并写入表头（文件不存在时执行）"""
    if file_path.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "List of Failed Cases"
    headers = ["CaseNodeID", "DateTime", "CaseLogPath", "FailureInfo"]
    ws.append(headers)
    # 列宽
    widths = [60, 20, 80, 100]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = w
    wb.save(file_path)


def append_failed_case_to_excel(file_path: Path, data: dict):
    """实时追加一行失败记录"""
    if not file_path.exists():
        init_excel_file(file_path)
    # 打开文件追加
    wb = load_workbook(file_path)
    ws = wb.active
    row = [
        data["nodeid"],
        data["execute_time"],
        data["case_log_path"],
        data["error_info"]
    ]
    ws.append(row)
    wb.save(file_path)
    wb.close()


@pytest.hookimpl(tryfirst=True, hookwrapper=True)
def pytest_runtest_makereport(item, call):
    outcome = yield
    rep = outcome.get_result()
    setattr(item, "rep_" + rep.when, rep)

    global excel_save_path
    if excel_save_path is None:
        return

    # 只捕获call阶段失败
    if rep.when == "call" and rep.failed:
        case_nodeid = item.nodeid
        error_msg = str(rep.longrepr) if rep.longrepr else "未知异常"
        exec_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        case_dir = ""
        if hasattr(item, "funcargs") and "case_log_dir" in item.funcargs:
            case_dir = str(item.funcargs["case_log_dir"])

        record = {
            "nodeid": case_nodeid,
            "execute_time": exec_time,
            "case_log_path": case_dir,
            "error_info": error_msg[:1500]
        }
        # ✅ 实时写入
        append_failed_case_to_excel(excel_save_path, record)
        print(f"[实时记录] 失败用例已写入Excel：{case_nodeid}")


@pytest.fixture(scope="session", autouse=True)
def setup_excel_path(dir_session):
    """when session started, init the excel path."""
    global excel_save_path
    excel_save_path = dir_session / "failed_cases_list.xlsx"
    init_excel_file(excel_save_path)
    print(f"Failed case excel path：{excel_save_path}")
    yield
    # the failed case is realtime added.
    
    